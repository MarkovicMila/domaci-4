import psycopg2 as psycopg2
import openpyxl as openpyxl

def query(q):
    con=psycopg2.connect(
        database='knjizara',
        port='5432',
        user='postgres',
        password='itoip',
        host='localhost'
    )
    cursor=con.cursor()
    cursor.execute(q)
    r=cursor.fetchall()
    cursor.close()
    con.close()
    return r


def select_all():
    l=query('SELECT * FROM knjiga')
    wb=openpyxl.Workbook()
    ws=wb.active
    ws.title='Sve knjige'
    ws['A1']='ID knjige'
    ws['B1']='Naziv'
    ws['C1']='Broj strana'
    ws['D1']='Cena'
    for i in range(2,len(l)+2):
        ws.cell(row=i,column=1).value=l[i-2][0]
        ws.cell(row=i,column=2).value=l[i-2][1]
        ws.cell(row=i,column=3).value=l[i-2][2]
        ws.cell(row=i,column=4).value=l[i-2][3]
    wb.save(filename='Sve knjige.xlsx')

def naziv_cena():
    l=query('SELECT naziv,cena FROM knjiga')
    wb=openpyxl.Workbook()
    ws=wb.active
    ws.title='Naziv_cena'
    ws['A1']='Naziv'
    ws['B1']='Cena'
    for i in range(2,len(l)+2):
        ws.cell(row=i,column=1).value=l[i-2][0]
        ws.cell(row=i,column=2).value=l[i-2][1]
    wb.save(filename='Naziv_cena.xlsx')

def prosecna_cena():
    l=query('SELECT cena FROM knjiga')
    sum=0
    for i in range(len(l)):
        for j in range(len(l[i])):
            sum+=l[i][j]
    print('Prosecna cena knjige je: ',sum/len(l))





print('''
1.Prikazati sve knjige sa svim informacijama
2.Prikazati nazive knjiga sa cenama
3.Prikazati prosecnu cenu knjiga
''')
opcija=eval(input())




if opcija==1:
    select_all()
elif opcija==2:
    naziv_cena()
elif opcija==3:
    prosecna_cena()
else:
    print('Neadekvatan izbor opcije! ')